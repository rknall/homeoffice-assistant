# SPDX-FileCopyrightText: 2025 Roland Knall <rknall@gmail.com>
# SPDX-License-Identifier: GPL-2.0-only
"""Expense report generator service."""

import io
import zipfile
from datetime import datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from slugify import slugify
from sqlalchemy.orm import Session

from src.integrations.base import DocumentProvider
from src.models import Event, Expense
from src.services import expense_service, integration_service


def _slugify_filename(name: str, max_length: int = 50) -> str:
    """Create a slug suitable for filenames."""
    slug = slugify(name, lowercase=True, separator="_")
    return slug[:max_length]


def _format_date(d: Any) -> str:
    """Format a date for filenames."""
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


class ExpenseReportGenerator:
    """Generator for expense reports with Excel and document packaging."""

    def __init__(
        self,
        db: Session,
        paperless: DocumentProvider | None = None,
    ) -> None:
        """Initialize the expense report generator.

        Args:
            db: Database session for querying expenses.
            paperless: Optional Paperless provider for downloading documents.
        """
        self.db = db
        self.paperless = paperless

    async def get_preview(self, event: Event) -> dict[str, Any]:
        """Return summary without generating files."""
        # Get base currency from company
        base_currency = event.company.base_currency if event.company else "EUR"

        expenses = expense_service.get_expenses(self.db, event.id)

        # Auto-convert any expenses missing conversion data
        await expense_service.ensure_expense_conversions(
            self.db, expenses, base_currency
        )

        # Now all expenses should have converted_amount (unless conversion failed)
        total = sum(
            e.converted_amount if e.converted_amount is not None else e.amount
            for e in expenses
        )
        documents_available = sum(1 for e in expenses if e.paperless_doc_id)

        by_category: dict[str, Decimal] = {}
        by_payment_type: dict[str, Decimal] = {}

        # Track currencies that were converted with their rates
        conversion_rates: dict[str, Decimal] = {}

        for expense in expenses:
            # Use converted amount for aggregations
            amount = (
                expense.converted_amount
                if expense.converted_amount is not None
                else expense.amount
            )
            cat = expense.category.value
            by_category[cat] = by_category.get(cat, Decimal(0)) + amount

            pt = expense.payment_type.value
            by_payment_type[pt] = by_payment_type.get(pt, Decimal(0)) + amount

            # Track conversion rates for display
            if (expense.currency.upper() != base_currency.upper()) and (
                expense.exchange_rate is not None
            ):
                conversion_rates[expense.currency.upper()] = expense.exchange_rate

        return {
            "event_id": event.id,
            "event_name": event.name,
            "company_name": event.company.name if event.company else None,
            "start_date": _format_date(event.start_date),
            "end_date": _format_date(event.end_date),
            "expense_count": len(expenses),
            "documents_available": documents_available,
            "total": float(total),
            "currency": base_currency,
            "by_category": {k: float(v) for k, v in by_category.items()},
            "by_payment_type": {k: float(v) for k, v in by_payment_type.items()},
            "paperless_configured": self.paperless is not None,
            "conversion_rates": (
                {k: float(v) for k, v in conversion_rates.items()}
                if conversion_rates
                else None
            ),
        }

    def _create_excel(
        self,
        event: Event,
        expenses: list[Expense],
    ) -> bytes:
        """Create Excel spreadsheet for expenses with currency conversion info."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Get base currency from company
        base_currency = event.company.base_currency if event.company else "EUR"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        amount_format = "#,##0.00"
        rate_format = "0.000000"
        date_format = "YYYY-MM-DD"

        # Title row
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value = f"Expense Report: {event.name}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        # Info rows
        ws["A2"] = f"Company: {event.company.name if event.company else 'N/A'}"
        start = _format_date(event.start_date)
        end = _format_date(event.end_date)
        ws["A3"] = f"Period: {start} to {end}"
        ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A5"] = f"Base Currency: {base_currency}"

        # Headers - expanded for currency conversion
        headers = [
            "#",
            "Date",
            "Description",
            "Category",
            "Payment",
            "Amount",
            "Currency",
            "Converted",
            "Base Curr",
            "Rate",
            "Rate Date",
            "Document",
        ]
        header_row = 7
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Data rows
        total_original = Decimal(0)
        total_converted = Decimal(0)
        for idx, expense in enumerate(expenses, 1):
            row = header_row + idx
            ws.cell(row=row, column=1, value=idx).border = border

            date_cell = ws.cell(row=row, column=2, value=expense.date)
            date_cell.number_format = date_format
            date_cell.border = border

            ws.cell(row=row, column=3, value=expense.description or "").border = border
            ws.cell(row=row, column=4, value=expense.category.value).border = border
            ws.cell(row=row, column=5, value=expense.payment_type.value).border = border

            # Original amount and currency
            amount_cell = ws.cell(row=row, column=6, value=float(expense.amount))
            amount_cell.number_format = amount_format
            amount_cell.border = border
            ws.cell(row=row, column=7, value=expense.currency).border = border

            # Converted amount (or original if same currency)
            converted = (
                expense.converted_amount
                if expense.converted_amount is not None
                else expense.amount
            )
            conv_cell = ws.cell(row=row, column=8, value=float(converted))
            conv_cell.number_format = amount_format
            conv_cell.border = border

            ws.cell(row=row, column=9, value=base_currency).border = border

            # Exchange rate
            rate = expense.exchange_rate if expense.exchange_rate else Decimal(1)
            rate_cell = ws.cell(row=row, column=10, value=float(rate))
            rate_cell.number_format = rate_format
            rate_cell.border = border

            # Rate date
            rate_date = expense.rate_date if expense.rate_date else expense.date
            rate_date_cell = ws.cell(row=row, column=11, value=rate_date)
            rate_date_cell.number_format = date_format
            rate_date_cell.border = border

            doc_ref = f"{idx:02d}_*.pdf" if expense.paperless_doc_id else "N/A"
            ws.cell(row=row, column=12, value=doc_ref).border = border

            total_original += expense.amount
            total_converted += converted

        # Total row
        total_row = header_row + len(expenses) + 1
        ws.cell(row=total_row, column=5, value="Totals:").font = Font(bold=True)

        orig_total_cell = ws.cell(row=total_row, column=6, value=float(total_original))
        orig_total_cell.font = Font(bold=True)
        orig_total_cell.number_format = amount_format

        conv_total_cell = ws.cell(row=total_row, column=8, value=float(total_converted))
        conv_total_cell.font = Font(bold=True)
        conv_total_cell.number_format = amount_format

        ws.cell(row=total_row, column=9, value=base_currency).font = Font(bold=True)

        # Adjust column widths
        column_widths = [5, 12, 35, 14, 12, 12, 8, 12, 8, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    async def generate(
        self,
        event: Event,
        expense_ids: list | None = None,
    ) -> tuple[bytes, list[Expense]]:
        """Generate ZIP with Excel and documents.

        Args:
            event: The event to generate a report for.
            expense_ids: Optional list of expense IDs to include. If None, includes all.

        Returns:
            Tuple of (zip_bytes, included_expenses)
        """
        if expense_ids:
            # Get specific expenses
            expenses = expense_service.get_expenses_by_ids(self.db, expense_ids)
            # Filter to only expenses for this event and exclude private expenses
            expenses = [
                e for e in expenses if e.event_id == event.id and not e.is_private
            ]
        else:
            # Get all non-private expenses for the event
            expenses = [
                e
                for e in expense_service.get_expenses(self.db, event.id)
                if not e.is_private
            ]

        expenses.sort(key=lambda e: e.date)

        # Create the Excel file
        excel_bytes = self._create_excel(event, expenses)

        # Create ZIP file
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            # Add Excel file
            event_slug = _slugify_filename(event.name)
            date_str = datetime.now().strftime("%Y-%m-%d")
            excel_name = f"expense_report_{event_slug}_{date_str}.xlsx"
            zip_file.writestr(excel_name, excel_bytes)

            # Add documents from Paperless if available
            if self.paperless:
                for idx, expense in enumerate(expenses, 1):
                    if expense.paperless_doc_id:
                        try:
                            (
                                content,
                                original_name,
                                _mime_type,
                            ) = await self.paperless.download_document(
                                expense.paperless_doc_id
                            )
                            # Extract extension from original filename or mime type
                            ext = "pdf"
                            if "." in original_name:
                                ext = original_name.rsplit(".", 1)[-1].lower()

                            # Create standardized filename
                            desc_slug = _slugify_filename(
                                expense.description or "document", 30
                            )
                            date_fmt = _format_date(expense.date)
                            new_filename = f"{idx:02d}_{date_fmt}_{desc_slug}.{ext}"

                            zip_file.writestr(f"documents/{new_filename}", content)
                        except Exception:  # noqa: S110
                            # Skip documents that fail to download
                            pass

        zip_buffer.seek(0)
        return zip_buffer.getvalue(), expenses

    def get_filename(self, event: Event) -> str:
        """Get the filename for the ZIP file."""
        event_slug = _slugify_filename(event.name)
        date_str = datetime.now().strftime("%Y-%m-%d")
        return f"expense_report_{event_slug}_{date_str}.zip"


async def create_report_generator(
    db: Session,
    event: Event,
) -> ExpenseReportGenerator:
    """Create a report generator with optional Paperless provider."""
    paperless_config = integration_service.get_active_document_provider(db)
    paperless = None

    if paperless_config:
        paperless = integration_service.create_provider_instance(paperless_config)
        if not isinstance(paperless, DocumentProvider):
            paperless = None

    return ExpenseReportGenerator(db, paperless)
